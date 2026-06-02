#!/usr/bin/env python3
"""
build_excel.py — генерирует план внедрения Aqua Delivery (.xlsx)
Использование: python3 build_excel.py <data.json> <output.xlsx>

data.json должен содержать поле partner_type: "new" | "existing"
  new      → два листа: «Описание процессов» + «План внедрения» (5 этапов, новый партнёр)
  existing → два листа: «Описание процессов» + «План внедрения» (только 1С, действующий партнёр)
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def thin_border():
    thin = Side(style='thin', color='BFBFBF')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')


# Цвета
HEADER_FILL = make_fill('1F4E79')
STAGE_FILL  = make_fill('2E75B6')
BLOC_FILLS  = {
    'Логистика':        make_fill('D6E4F0'),
    'Операторы / CRM':  make_fill('E8F5E9'),
    'Склад':            make_fill('FFF9C4'),
    'Бухгалтерия / 1С': make_fill('FCE4EC'),
    'Маркировка':       make_fill('F3E5F5'),
    'Аналитика':        make_fill('E0F7FA'),
}
DEFAULT_BLOC_FILL = make_fill('F5F5F5')
ROW_ALT_FILL      = make_fill('FAFAFA')
WHITE_FILL        = make_fill('FFFFFF')
FOOTER_FILL       = make_fill('FFF3CD')
TOTAL_FILL        = make_fill('D9EAD3')

WHITE_BOLD  = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
DARK_BOLD   = Font(name='Calibri', bold=True, color='1F2D3D', size=10)
DARK_NORMAL = Font(name='Calibri', color='1F2D3D', size=10)
RED_BOLD    = Font(name='Calibri', bold=True, color='C00000', size=11)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical='top')
CENTER_MID  = Alignment(horizontal='center', vertical='center', wrap_text=True)


# ─────────────────────────────────────────────────────────────────────────────
# Лист 1: Описание процессов (одинаков для обоих типов)
# ─────────────────────────────────────────────────────────────────────────────

def build_processes_sheet(ws, processes):
    ws.title = 'Описание процессов'

    headers = [
        'Блок', 'Процесс', 'Было\n(текущий процесс)', 'Стало\n(после внедрения)',
        'Соответствует готовому\nпродукту без доработок',
        'Инструмент\nAqua Delivery',
        'Требует доработки /\nрешения',
        'Результат\nулучшения процесса',
        'Риски / комментарии', 'Согласовано',
    ]
    col_widths = [20, 40, 45, 45, 18, 25, 30, 30, 40, 14]

    ws.row_dimensions[1].height = 40
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER_MID
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    row_num = 2
    for p in processes:
        bloc = p.get('bloc', '')
        bloc_fill = BLOC_FILLS.get(bloc, DEFAULT_BLOC_FILL)
        alt = (row_num % 2 == 0)

        values = [
            bloc,
            p.get('process', ''),
            p.get('bylo', ''),
            p.get('stalo', ''),
            '',
            p.get('instrument', ''),
            '',
            '',
            p.get('riski', ''),
            '',
        ]

        ws.row_dimensions[row_num].height = max(
            60, max(len(str(v)) // 45 * 15 + 15 for v in values if v)
        )

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.alignment = WRAP_ALIGN
            cell.border = thin_border()
            if col_idx in (1, 6):
                cell.fill = bloc_fill
                cell.font = DARK_BOLD
            else:
                cell.fill = ROW_ALT_FILL if alt else WHITE_FILL
                cell.font = DARK_NORMAL

        row_num += 1

    ws.freeze_panes = 'A2'


# ─────────────────────────────────────────────────────────────────────────────
# Лист 2 — вариант A: Новый партнёр (5 этапов)
# ─────────────────────────────────────────────────────────────────────────────

def build_plan_sheet_new(ws, plan_data):
    ws.title = 'План внедрения'

    partner = plan_data.get('partner', {})
    partner_name = partner.get('name', '')
    end_date = partner.get('end_date', '')

    # Заголовок документа
    ws.merge_cells('A1:F1')
    title_cell = ws.cell(row=1, column=1, value=f'ПЛАН ВНЕДРЕНИЯ — {partner_name}')
    title_cell.fill = HEADER_FILL
    title_cell.font = WHITE_BOLD
    title_cell.alignment = CENTER_MID
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:F2')
    subtitle_cell = ws.cell(row=2, column=1, value=f'Дата завершения: {end_date}')
    subtitle_cell.fill = make_fill('2E75B6')
    subtitle_cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    subtitle_cell.alignment = CENTER_MID
    ws.row_dimensions[2].height = 20

    headers = ['№', 'Задача', 'Описание процесса', 'Зона ответственности партнёра', 'Срок', 'Стоимость']
    col_widths = [5, 40, 55, 40, 14, 30]

    ws.row_dimensions[3].height = 35
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.fill = STAGE_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER_MID
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    row_num = 4
    step_counter = 1

    for stage_data in plan_data.get('stages', []):
        # Строка этапа
        ws.merge_cells(f'A{row_num}:F{row_num}')
        stage_cell = ws.cell(row=row_num, column=1, value=stage_data.get('stage', ''))
        stage_cell.fill = STAGE_FILL
        stage_cell.font = WHITE_BOLD
        stage_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        stage_cell.border = thin_border()
        ws.row_dimensions[row_num].height = 25
        row_num += 1

        for step in stage_data.get('steps', []):
            alt = (step_counter % 2 == 0)
            row_fill = ROW_ALT_FILL if alt else WHITE_FILL

            values = [
                step_counter,
                step.get('name', ''),
                step.get('process', ''),
                step.get('partner_responsibility', ''),
                step.get('expected_date', ''),
                step.get('cost', ''),
            ]

            max_len = max(len(str(v)) for v in values if v)
            ws.row_dimensions[row_num].height = max(50, max_len // 50 * 15 + 30)

            for col_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.fill = row_fill
                cell.font = DARK_NORMAL
                cell.alignment = WRAP_ALIGN
                cell.border = thin_border()

            step_counter += 1
            row_num += 1

    # Итог
    ws.merge_cells(f'A{row_num}:E{row_num}')
    total_label = ws.cell(row=row_num, column=1, value='ИТОГО:')
    total_label.fill = TOTAL_FILL
    total_label.font = DARK_BOLD
    total_label.alignment = Alignment(horizontal='right', vertical='center')
    total_label.border = thin_border()

    total_val = ws.cell(row=row_num, column=6, value=plan_data.get('total_cost', 'УТОЧНИТЬ'))
    total_val.fill = TOTAL_FILL
    total_val.font = DARK_BOLD
    total_val.alignment = CENTER_MID
    total_val.border = thin_border()

    ws.freeze_panes = 'A4'


# ─────────────────────────────────────────────────────────────────────────────
# Лист 2 — вариант B: Действующий партнёр (только 1С)
# ─────────────────────────────────────────────────────────────────────────────

def build_plan_sheet_existing(ws, plan_data):
    ws.title = 'План внедрения'

    partner = plan_data.get('partner', {})
    partner_name = partner.get('name', '')
    end_date = partner.get('end_date', '')

    ws.merge_cells('A1:E1')
    title_cell = ws.cell(row=1, column=1, value=f'ПЛАН ВНЕДРЕНИЯ 1С — {partner_name}')
    title_cell.fill = HEADER_FILL
    title_cell.font = WHITE_BOLD
    title_cell.alignment = CENTER_MID
    ws.row_dimensions[1].height = 30

    ws.merge_cells('A2:E2')
    subtitle_cell = ws.cell(row=2, column=1, value=f'Дата завершения: {end_date}')
    subtitle_cell.fill = make_fill('2E75B6')
    subtitle_cell.font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    subtitle_cell.alignment = CENTER_MID
    ws.row_dimensions[2].height = 20

    headers = ['№', 'Задача', 'Срок', 'Стоимость', 'Примечание']
    col_widths = [5, 80, 14, 30, 30]

    ws.row_dimensions[3].height = 35
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.fill = STAGE_FILL
        cell.font = WHITE_BOLD
        cell.alignment = CENTER_MID
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    row_num = 4
    for step in plan_data.get('steps', []):
        num = step.get('num', '')
        name = step.get('name', '')
        date = step.get('expected_date', '')
        cost = step.get('cost', '')
        note = step.get('note', '')

        alt = (row_num % 2 == 0)
        row_fill = ROW_ALT_FILL if alt else WHITE_FILL

        ws.row_dimensions[row_num].height = max(50, len(name) // 80 * 15 + 30)

        for col_idx, val in enumerate([num, name, date, cost, note], 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.fill = row_fill
            cell.font = DARK_NORMAL
            cell.alignment = WRAP_ALIGN
            cell.border = thin_border()

        row_num += 1

    # Footer
    row_num += 1
    for entry in plan_data.get('footer', []):
        ws.merge_cells(f'A{row_num}:D{row_num}')
        label = ws.cell(
            row=row_num, column=1,
            value=f'{entry.get("marker", "")} {entry.get("name", "")}'
        )
        label.fill = FOOTER_FILL
        label.font = DARK_BOLD
        label.alignment = WRAP_ALIGN
        label.border = thin_border()

        cost_cell = ws.cell(row=row_num, column=5, value=entry.get('cost', ''))
        cost_cell.fill = FOOTER_FILL
        cost_cell.font = DARK_BOLD
        cost_cell.alignment = CENTER_MID
        cost_cell.border = thin_border()

        ws.row_dimensions[row_num].height = 25
        row_num += 1

    # Итог
    ws.merge_cells(f'A{row_num}:D{row_num}')
    total_label = ws.cell(row=row_num, column=1, value='ИТОГО:')
    total_label.fill = TOTAL_FILL
    total_label.font = DARK_BOLD
    total_label.alignment = Alignment(horizontal='right', vertical='center')
    total_label.border = thin_border()

    total_val = ws.cell(row=row_num, column=5, value=plan_data.get('total_cost', 'УТОЧНИТЬ'))
    total_val.fill = TOTAL_FILL
    total_val.font = DARK_BOLD
    total_val.alignment = CENTER_MID
    total_val.border = thin_border()

    ws.freeze_panes = 'A4'


# ─────────────────────────────────────────────────────────────────────────────
# Точка входа
# ─────────────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print('Usage: python3 build_excel.py <data.json> <output.xlsx>')
        sys.exit(1)

    data_path   = sys.argv[1]
    output_path = sys.argv[2]

    with open(data_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    partner_type = data.get('partner_type', 'new')
    processes    = data.get('processes', [])
    plan         = data.get('plan', {})

    wb = Workbook()
    ws_processes = wb.active
    build_processes_sheet(ws_processes, processes)

    ws_plan = wb.create_sheet()
    if partner_type == 'existing':
        build_plan_sheet_existing(ws_plan, plan)
    else:
        build_plan_sheet_new(ws_plan, plan)

    wb.save(output_path)
    print(f'✅ Файл сохранён: {output_path}')


if __name__ == '__main__':
    main()
